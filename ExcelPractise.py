
# import load_workbook
from openpyxl import load_workbook
import pandas as pd


import xlrd
# set file path
filepath="practise.xlsx"

# WIll read initially but not after one run
"""dd = xlrd.open_workbook(filepath)
Team = dd.sheet_by_name("Team")
print("Here",Team.cell_value(1, 1))"""


tmp_name = "Team"

# load demo.xlsx
wb=load_workbook(filepath)
wb2 = load_workbook(filepath,data_only=True)
#df = pd.read_excel(filepath,sheet_name=tmp_name)
#print(df)

# get Sheet
print(wb.sheetnames)
sheet=wb[tmp_name]
sheet2=wb2[tmp_name]


print(sheet2.cell(row=2, column=2).value)
# set value for cell B2=2
sheet.cell(row=2, column=3).value = 2


""""# copy sheet
target=wb.copy_worksheet(source)

wb.remove(wb[tmp_name])
print(target.title)
target.title = tmp_name
print(target.title)"""
# save workbook
wb.save(filepath)