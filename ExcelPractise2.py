import xlrd
import xlwt
# set file path

# import load_workbook
from openpyxl import load_workbook
import pandas as pd


import xlrd
# set file path
filepath="practise.xlsx"

# WIll read initially but not after one run
"""dd = xlrd.open_workbook(filepath)
Team = dd.sheet_by_name("Team")
print("Here",Team.cell_value(1, 1))"""


tmp_name = "Team"

# load demo.xlsx
wb=load_workbook(filepath)
wbd = load_workbook(filename=filepath, data_only=True)

# get Sheet "Team"
sheetd=wbd[tmp_name]
sheet=wb[tmp_name]
print("Data",sheetd.cell(row=2, column=1).value)
print("Data",sheetd.cell(row=2, column=2).value)
print("Data",sheetd.cell(row=2, column=3).value)

formula = sheet.cell(row=2, column=2).value
print(sheet.cell(row=2, column=1).value)
print(sheet.cell(row=2, column=2).value)
print(sheet.cell(row=2, column=3).value)


# copy sheet
target=wb.copy_worksheet(sheet)
# set value for cell B2=2
target.cell(row=2, column=3).value = "2"
wb.save(filepath)
target.cell(row=2, column=2).value = str(formula)


wb.remove(wb[tmp_name])
print(target.title)
target.title = tmp_name
print(target.title)
# save workbook
wb.save(filepath)

